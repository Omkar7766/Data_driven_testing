import openpyxl

file="C:/Users/LENOVO/Documents/Omkar/testdata.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook['Sheet1']

#noof_rows=sheet.max_row
#columns=sheet.max_column

#print("no of rows:",noof_rows)
#print("no of columns:",columns)


sheet.cell(1,1).value="SID"
sheet.cell(1,2).value="NAME"
sheet.cell(1,3).value="ROLE"

sheet.cell(2,1).value=567
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=567
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="developer"

workbook.save(file)

print("Successfully written data into excel")

